import requests
from openpyxl.worksheet.worksheet import Worksheet

SHEET_CHANGE_VALUE = ' SHEET UPDATE'
HEADER_ROW_NUMBER = 4
VALUE_ROW_NUMBER = 6


class BulkUpdateManager:

    def __init__(self, ingest_api, ingest_url):
        self.ingest_api = ingest_api
        self.ingest_url = ingest_url

    @staticmethod
    def get_entities_by_submission_id_and_type(self, submission_id, entity_type):
        response = requests.get(self.ingest_url + f'/submissionEnvelopes/{submission_id}/{entity_type}').json()
        return response.get('_embedded').get(entity_type)

    @staticmethod
    def get_id_from_entity(entity):
        return entity['_links']['self']['href'].split('/')[-1]

    def update_content(self, entity_type, entity_id, original_content):
        self.ingest_api.patch(self.ingest_url + f'/{entity_type}/' + entity_id, {'content': original_content})

    def update_project_title(self, project_sheet):
        project_title_column_index = self.__get_cell_by_header_name(project_sheet, 'project_title')
        orig_project_title = project_sheet.cell(VALUE_ROW_NUMBER, project_title_column_index).value
        updated_project_title = orig_project_title + SHEET_CHANGE_VALUE
        project_sheet.cell(VALUE_ROW_NUMBER, project_title_column_index, updated_project_title)

        return updated_project_title

    def update_biomaterial_name(self, specimen_sheet):
        biomaterial_name_column_index = self.__get_cell_by_header_name(specimen_sheet, 'biomaterial_name')
        orig_biomaterial_name = specimen_sheet.cell(VALUE_ROW_NUMBER, biomaterial_name_column_index).value
        updated_biomaterial_name = orig_biomaterial_name + SHEET_CHANGE_VALUE
        specimen_sheet.cell(VALUE_ROW_NUMBER, biomaterial_name_column_index, updated_biomaterial_name)

        modified_biomaterial_id = self.__get_entity_id_by_sheet_and_entity_type(specimen_sheet, 'biomaterials')

        return updated_biomaterial_name, modified_biomaterial_id

    @staticmethod
    def save_modified_spreadsheet(path_to_spreadsheet, spreadsheet):
        spreadsheet.save(path_to_spreadsheet)

    @staticmethod
    def __get_cell_by_header_name(worksheet: Worksheet, header_name):
        header_row = worksheet[HEADER_ROW_NUMBER]
        for cell in header_row:
            if header_name in cell.value:
                return cell.col_idx
        return None

    def __get_entity_id_by_sheet_and_entity_type(self, sheet, entity_type):
        uuid_column_index = self.__get_cell_by_header_name(sheet, 'uuid')
        entity_uuid = sheet.cell(VALUE_ROW_NUMBER, uuid_column_index).value
        modified_entity = self.ingest_api.get_entity_by_uuid(entity_type, entity_uuid)
        return self.__get_id_from_entity(modified_entity)

    @staticmethod
    def __get_id_from_entity(entity):
        return entity['_links']['self']['href'].split('/')[-1]
